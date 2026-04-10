from openpyxl import load_workbook
import requests
from datetime import datetime
import pandas as pd

# Configurações — ajuste conforme o ambiente
CAMINHO_PLANILHA = "caminho/para/DDR.xlsx"
ABA_PLANILHA     = "Arquivos 2025"
COLUNA_DATAS     = "B"
COLUNA_COTACOES  = "D"


def buscar_cotacao_dolar(data_base):
    """
    Consulta a API PTAX do BACEN e retorna a cotação de venda do dólar
    para a data informada.

    Args:
        data_base (datetime): data de referência

    Returns:
        float | None: cotação de venda ou None se não encontrada
    """
    data_formatada = data_base.strftime('%m-%d-%Y')
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
        f"CotacaoDolarDia(dataCotacao=@dataCotacao)"
        f"?@dataCotacao=%27{data_formatada}%27&$format=json"
    )

    response = requests.get(url)
    if response.status_code == 200:
        dados = response.json()
        if "value" in dados and len(dados["value"]) > 0:
            return dados["value"][0]["cotacaoVenda"]
        else:
            print(f"Cotacao nao encontrada para {data_formatada}")
    else:
        print(f"Erro na requisicao: {response.status_code}")

    return None


def atualizar_planilha():
    """
    Percorre a planilha linha por linha e preenche a cotação do dólar
    nas células ainda vazias, consultando a API PTAX para cada data.
    """
    wb = load_workbook(filename=CAMINHO_PLANILHA)
    ws = wb[ABA_PLANILHA]

    for row in range(2, ws.max_row + 1):
        celula_data    = ws[f"{COLUNA_DATAS}{row}"]
        celula_cotacao = ws[f"{COLUNA_COTACOES}{row}"]

        if celula_data.value:
            if celula_cotacao.value is None:
                data_base = pd.to_datetime(celula_data.value, dayfirst=True)
                cotacao = buscar_cotacao_dolar(data_base)
                celula_cotacao.value = cotacao
        else:
            print(f"Linha {row} ignorada (data-base vazia)")

    wb.save(CAMINHO_PLANILHA)
    print("Atualizacao concluida!")


if __name__ == "__main__":
    atualizar_planilha()
