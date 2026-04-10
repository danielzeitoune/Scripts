from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from datetime import datetime, timedelta, date
import time
from openpyxl import load_workbook
import pandas as pd
import os

# Configurações — ajuste conforme o ambiente
CAMINHO_PLANILHA  = "caminho/para/DDR.xlsx"
CAMINHO_FERIADOS  = "caminho/para/feriados_nacionais.xlsx"
ABA_PLANILHA      = "Arquivos 2025"
COLUNA_DATAS      = "B"
COLUNA_MULTIPLICADOR = "I"
CODIGO_SERIE_BCB  = "13804"

# Corrige erro de SSL no webdriver-manager
os.environ['WDM_SSL_VERIFY'] = '0'


def obter_multiplicador():
    """
    Acessa o SGS (Sistema Gerenciador de Series Temporais) do BACEN via Selenium,
    consulta a serie 13804 e retorna o ultimo valor disponivel multiplicado por 100.

    Returns:
        float: multiplicador calculado
    """
    servico   = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico)
    navegador.get(
        "https://www3.bcb.gov.br/sgspub/localizarseries/localizarSeries.do"
        "?method=prepararTelaLocalizarSeries"
    )
    time.sleep(1)

    # Aceita alerta inicial da pagina
    navegador.switch_to.alert.accept()
    time.sleep(1)

    # Busca a serie pelo codigo
    navegador.find_element(By.XPATH, '//*[@id="txCodigo"]').send_keys(
        CODIGO_SERIE_BCB, Keys.ENTER
    )
    time.sleep(1)

    # Seleciona todas e consulta
    navegador.find_element(By.XPATH, '//*[@id="botaoMarcar"]/input').click()
    navegador.find_element(
        By.XPATH,
        '/html/body/form/center/span/center/table/tbody/tr/td[4]/div/input'
    ).click()
    time.sleep(1)

    # Calcula o range de datas (3 dias uteis anteriores a hoje)
    df_feriados   = pd.read_excel(CAMINHO_FERIADOS, engine='openpyxl')
    lista_feriados = pd.to_datetime(df_feriados['Data']).dt.date.tolist()

    data_final   = date.today() - timedelta(days=1)
    data_inicial = data_final
    lista_dias   = []
    dias_uteis   = 0

    while dias_uteis < 3:
        if data_inicial.weekday() < 5 and data_inicial not in lista_feriados:
            lista_dias.append(data_inicial)
            dias_uteis += 1
        data_inicial -= timedelta(days=1)

    lista_dias.sort()

    data_final_fmt   = datetime.strftime(data_final,    '%d%m%Y')
    data_inicial_fmt = datetime.strftime(lista_dias[0], '%d%m%Y')

    # Preenche o range de datas e consulta
    navegador.find_element(By.XPATH, '//*[@id="dataFim"]').clear()
    navegador.find_element(By.XPATH, '//*[@id="dataFim"]').send_keys(data_final_fmt)
    navegador.find_element(By.XPATH, '//*[@id="dataInicio"]').clear()
    navegador.find_element(By.XPATH, '//*[@id="dataInicio"]').send_keys(data_inicial_fmt)

    navegador.find_element(
        By.XPATH, '/html/body/center/form/div[2]/input[2]'
    ).click()
    time.sleep(1)

    # Captura o ultimo valor da tabela
    ultimo_valor = navegador.find_element(
        By.XPATH, '//*[@id="valoresSeries"]/tbody/tr[6]/td[2]/div/span'
    ).text

    navegador.quit()

    ultimo_valor_mult = float(ultimo_valor.replace(',', '.')) * 100
    print(f"Ultimo valor encontrado: {ultimo_valor}")
    print(f"Multiplicador calculado: {ultimo_valor_mult}")
    return ultimo_valor_mult


def atualizar_planilha(ultimo_valor_mult):
    """
    Grava o multiplicador na primeira celula vazia da coluna correspondente.

    Args:
        ultimo_valor_mult (float): valor a ser inserido na planilha
    """
    wb = load_workbook(filename=CAMINHO_PLANILHA)
    ws = wb[ABA_PLANILHA]

    for row in range(2, ws.max_row + 1):
        celula_data   = ws[f"{COLUNA_DATAS}{row}"]
        celula_multip = ws[f"{COLUNA_MULTIPLICADOR}{row}"]

        if celula_data.value and isinstance(celula_data.value, (datetime, str)):
            try:
                if isinstance(celula_data.value, str):
                    datetime.strptime(celula_data.value, "%d/%m/%Y")
                if celula_multip.value is None:
                    celula_multip.value = ultimo_valor_mult
                    break
            except ValueError:
                print(f"Linha {row} ignorada (data invalida: '{celula_data.value}')")
        else:
            print(f"Linha {row} ignorada (data-base vazia ou invalida)")

    wb.save(CAMINHO_PLANILHA)
    print("Atualizacao concluida!")


if __name__ == "__main__":
    multiplicador = obter_multiplicador()
    atualizar_planilha(multiplicador)
